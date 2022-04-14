from openpyxl import load_workbook
from openpyxl.styles import Font


def couleurCellule():
	"""
	Fonction qui permet de changer la couleur de la cellule
	"""
	# Pour chaque cellule
	for typeCell in sheet.iter_cols(min_row=20, max_row=20, min_col=1, max_col=5):

		# Aucune connaissance des valeurs des index.
		if typeCell[0].value is not None:
			# Valeur RGB en fonction de la valeur de la cellule
			rgb = (int(255 - (typeCell[0].value * 2.55)), int(typeCell[0].value * 2.55), 0)
			# Converti en hexadecimal
			hexa = '%02x%02x%02x' % rgb
			# Couleur de la cellule
			typeCell[0].value = Font(color=hexa)


def calculMoyenne(colonneactuelle):
	# Prend la dernière valeur de la colonne
	row = 2
	precis = dicoCours[cours][0] + str(row)
	while sheet[precis].value is not None:
		row += 1
		precis = dicoCours[cours][0] + str(row)

	# Fais la somme de chaque valeur dans la colonne.
	somme = 0
	for i in sheet.iter_rows(min_row=2, max_row=row-1, min_col=colonneactuelle[1], max_col=colonneactuelle[1]):
		somme += i[0].value * 5

	# Ecrit la moyenne des notes en bas de la colonne.
	sheet[colonneactuelle[0] + "20"] = somme / (row - 1)


# Charge le document
path = "C:\\Users\\diama\\OneDrive\\TEP 1\\Note.xlsx"
wb = load_workbook(path)
sheet = wb.active

# Demande le cours pour la clé du dictionnaire.
cours = input("Quel cours: AO, FR, PY, SYS, GBD? ").lower()

# Dictionnaire pour les numéros des colonnes et leurs lettres associées.
dicoCours = {"ao": ("A", 1), "fr": ("B", 2), "py": ("C", 3), "sys": ("D", 4), "gbd": ("E", 5)}
# Créer boucle pour demander note plusieurs fois.
while True:
	# Nom des cours sur la première ligne. AO, FR, Prog, Sys Expl, GBDD#
	note = input("\nQ pour quitter" + "\nEntrez la note: ")
	# stop tout si user demande de quiter
	if note.lower() == "q":
		break

	# Split et converti sur 20 la note.
	note = note.split("/")
	noteSurVingt = int(note[0]) / int(note[1]) * 20

	# Cherche une case vide pour la colonne du cours.
	case = 2
	colonne = dicoCours[cours][0] + str(case)
	while sheet[colonne].value is not None:
		case += 1
		colonne = dicoCours[cours][0] + str(case)

	# Ecrit la note/20 sur l'emplacement libre
	sheet[colonne] = noteSurVingt

calculMoyenne(dicoCours[cours])
couleurCellule()
wb.save(path)
